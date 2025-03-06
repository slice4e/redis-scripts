import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference


fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")


def highlight_cells(ws, top_n, reverse=False):
    values = []
    for row in ws.iter_rows(min_row=6, min_col=3):
        for cell in row:
            try:
                cell_value = float(cell.value)
                values.append((cell_value, cell))
            except (TypeError, ValueError):
                continue
    values.sort(key=lambda x: x[0], reverse=reverse)
    for i in range(min(top_n, len(values))):
        values[i][1].fill = fill_green


def create_excel_for_top(df_max_rps, top):
    df_max_rps = df_max_rps[df_max_rps["top"] == top]
    df_total_time = df_max_rps.pivot(index=['parallel', 'ws_search'], columns=['data_type', 'graph_degree', 'num_threads', 'top', 'quantization', 'ws_construction'], values='total_time')
    df_mean_time = df_max_rps.pivot(index=['parallel', 'ws_search'], columns=['data_type', 'graph_degree', 'num_threads', 'top', 'quantization', 'ws_construction'], values='mean_time')
    df_precision = df_max_rps.pivot(index=['parallel', 'ws_search'], columns=['data_type', 'graph_degree', 'num_threads', 'top', 'quantization', 'ws_construction'], values='precision')
    df_rps = df_max_rps.pivot(index=['parallel', 'ws_search'], columns=['data_type', 'graph_degree', 'num_threads', 'top', 'quantization', 'ws_construction', ], values='rps')

    df_uploading_time = pd.DataFrame.from_dict(settings, orient='index')[['uploading_time']]
    df_used_memory = pd.DataFrame.from_dict(settings, orient='index')[['used_memory']]


    quantization_values = df_max_rps['quantization'].unique()
    parallel_values = df_max_rps['parallel'].unique()
    closest_precision_rps = []

    for quant in quantization_values:
        for parallel in parallel_values:
            df_quant_parallel = df_max_rps[(df_max_rps['quantization'] == quant) & (df_max_rps['parallel'] == parallel)]
            if not df_quant_parallel.empty:
                df_quant_parallel['precision_diff'] = abs(df_quant_parallel['precision'] - 0.95)
                closest_row = df_quant_parallel.loc[df_quant_parallel['precision_diff'].idxmin()]
                closest_precision_rps.append({
                    'quantization': quant,
                    'parallel': parallel,
                    'precision': closest_row['precision'],
                    'rps': closest_row['rps']
                })

    df_closest_precision_rps = pd.DataFrame(closest_precision_rps)
    df_rps_pivot = df_closest_precision_rps.pivot(index='parallel', columns='quantization', values='rps')
    df_precision_pivot = df_closest_precision_rps.pivot(index='parallel', columns='quantization', values='precision')

    excel_path = f"summary/results-top{top}.xlsx"
    with pd.ExcelWriter(excel_path) as writer:
        df_total_time.to_excel(writer, sheet_name="Total Time")
        df_mean_time.to_excel(writer, sheet_name="Mean Time")
        df_precision.to_excel(writer, sheet_name="Precision")
        df_rps.to_excel(writer, sheet_name="RPS")
        df_uploading_time.to_excel(writer, sheet_name="Uploading Time")
        df_used_memory.to_excel(writer, sheet_name="Used Memory")

        worksheet = writer.book.create_sheet("Precision 95")
        worksheet.title = "Precision 95"
        
        worksheet.cell(row=1, column=1, value="RPS")
        for r in dataframe_to_rows(df_rps_pivot, index=True, header=True):
            worksheet.append(r)
        
        start_row = df_rps_pivot.shape[0] + 5
        worksheet.cell(row=start_row, column=1, value="Precision")
        for r in dataframe_to_rows(df_precision_pivot, index=True, header=True):
            worksheet.append(r)

        # Add Bar chart to the sheet
        chart = BarChart()
        chart.title = "RPS @ ~95%"
        chart.x_axis.title = "Parallel"
        chart.y_axis.title = "RPS"

        data = Reference(worksheet, min_col=2, min_row=2, max_row=df_rps_pivot.shape[0] + 3, max_col=df_rps_pivot.shape[1] + 1)
        categories = Reference(worksheet, min_col=1, min_row=2, max_row=df_rps_pivot.shape[0] + 3)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4
        worksheet.add_chart(chart, "E2")


    wb = load_workbook(excel_path)


    for sheet_name in ["Total Time", "Mean Time"]:
        ws = wb[sheet_name]
        highlight_cells(ws, top_n=3, reverse=False)
        ws.column_dimensions['B'].width = 20

    for sheet_name in ["Precision", "RPS"]:
        ws = wb[sheet_name]
        highlight_cells(ws, top_n=3, reverse=True)
        ws.column_dimensions['B'].width = 20 

    for sheet_name in ["Uploading Time", "Used Memory"]:
        ws = wb[sheet_name]
        ws.column_dimensions['A'].width = 40 
        ws.column_dimensions['B'].width = 20
        min_value = None
        min_cell = None
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                try:
                    cell_value = float(cell.value)
                    if min_value is None or cell_value < min_value:
                        min_value = cell_value
                        min_cell = cell
                except (TypeError, ValueError):
                    continue
        if min_cell:
            min_cell.fill = fill_green

    wb.save(excel_path)


paths = []
path = os.getcwd()
files = [f for f in os.listdir(path) if (os.path.isfile(os.path.join(path, f))) and (f.endswith('.json')) ]
settings = {}
results = []
for file in files:
    if "upload"  in file:
        with open(os.path.join(path, file)) as f:
            print(file) 
            data = json.load(f)
            experiment_name = data["params"]["experiment"]
            configuration = {}
            configuration["num_threads"] = data["params"]["svs_config"]["NUM_THREADS"] 
            configuration["graph_degree"] = data["params"]["svs_config"]["GRAPH_DEGREE"] 
            configuration["ws_construction"] = data["params"]["svs_config"]["WS_CONSTRUCTION"] 
            configuration["quantization"] = data["params"]["svs_config"]["QUANTIZATION"]
            configuration["uploading_time"] = data["results"]["total_time"]
            configuration["used_memory"] = data["results"]["memory_usage"]["used_memory"][0]
            settings[experiment_name] = configuration

for file in files:
    if not "upload" in file:
        with open(os.path.join(path, file)) as f:
            print(file) 
            data = json.load(f)
            result = {}
            experiment_name = data["params"]["experiment"]
            result["parallel"] = data["params"]["parallel"]
            result["top"] = data["params"].get("top", 100)
            result["data_type"] = data["params"]["search_params"].get("data_type", "FLOAT32")
            result["total_time"] = data["results"]["total_time"]
            result["mean_time"] = data["results"]["mean_time"]
            result["precision"] = data["results"]["mean_precisions"]
            result["rps"] = data["results"]["rps"]
            result["ws_search"] = data["params"]["search_params"]["WS_SEARCH"]
            result["num_threads"] = settings[experiment_name]["num_threads"]
            result["graph_degree"] = settings[experiment_name]["graph_degree"]
            result["ws_construction"] = settings[experiment_name]["ws_construction"]
            result["quantization"] = settings[experiment_name]["quantization"]
            results.append(result)

result_json_path = os.path.join(path, "summary/results.json")
os.makedirs(os.path.dirname(result_json_path), exist_ok=True)
with open(result_json_path, "w") as f:
    json.dump(results, f)


df = pd.DataFrame(results)

idx = df.groupby(['parallel', 'ws_search', 'data_type', 'graph_degree', 'num_threads', 'top', 'quantization', 'ws_construction'])['rps'].idxmax()
df_max_rps = df.loc[idx]
create_excel_for_top(df_max_rps, 10)
#create_excel_for_top(df_max_rps, 100)

