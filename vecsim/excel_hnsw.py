import json
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")


def highlight_cells(ws, top_n, reverse=False):
    values = []
    for row in ws.iter_rows(min_row=6, min_col=3):
        for cell in row:
            try:
                cell_value = float(cell.value)
                values.append((cell_value, cell))
            except (TypeError, ValueError):
                continue
    values.sort(key=lambda x: x[0], reverse=reverse)
    for i in range(min(top_n, len(values))):
        values[i][1].fill = fill_green



def create_excel_for_top(df_max_rps, top):
    df_max_rps = df_max_rps[df_max_rps["top"] == top]
    # Create dataframe for each table
    df_total_time = df_max_rps.pivot(index=['parallel', 'ef_search'], columns=['top', 'M',  'ef_construction'], values='total_time')
    df_mean_time = df_max_rps.pivot(index=['parallel', 'ef_search'], columns=['top', 'M', 'ef_construction'], values='mean_time')
    df_precision = df_max_rps.pivot(index=['parallel', 'ef_search'], columns=['top', 'M', 'ef_construction'], values='precision')
    df_rps = df_max_rps.pivot(index=['parallel', 'ef_search'], columns=['top', 'M', 'ef_construction', ], values='rps')

    # Create DataFrame for used_mem and uploading_Time
    df_uploading_time = pd.DataFrame.from_dict(settings, orient='index')[['uploading_time']]
    df_used_memory = pd.DataFrame.from_dict(settings, orient='index')[['used_memory']]

    # Save to excel file
    excel_path = f"summary/results-hnsw-top{top}.xlsx"
    with pd.ExcelWriter(excel_path) as writer:
        df_total_time.to_excel(writer, sheet_name="Total Time")
        df_mean_time.to_excel(writer, sheet_name="Mean Time")
        df_precision.to_excel(writer, sheet_name="Precision")
        df_rps.to_excel(writer, sheet_name="RPS")
        df_uploading_time.to_excel(writer, sheet_name="Uploading Time")
        df_used_memory.to_excel(writer, sheet_name="Used Memory")

    wb = load_workbook(excel_path)

    for sheet_name in ["Total Time", "Mean Time"]:
        ws = wb[sheet_name]
        highlight_cells(ws, top_n=3, reverse=False)
        ws.column_dimensions['B'].width = 20

    for sheet_name in ["Precision", "RPS"]:
        ws = wb[sheet_name]
        highlight_cells(ws, top_n=3, reverse=True)
        ws.column_dimensions['B'].width = 20

    # Highlight lowest value in uploading time and used mem
    for sheet_name in ["Uploading Time", "Used Memory"]:
        ws = wb[sheet_name]
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        min_value = None
        min_cell = None
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                try:
                    cell_value = float(cell.value)
                    if min_value is None or cell_value < min_value:
                        min_value = cell_value
                        min_cell = cell
                except (TypeError, ValueError):
                    continue
        if min_cell:
            min_cell.fill = fill_green

    wb.save(excel_path)

paths = []
path = os.getcwd()
files = [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f))]
settings = {}
results = []
for file in files:
    if "upload" in file:
        with open(os.path.join(path, file)) as f:
            data = json.load(f)
            experiment_name = data["params"]["experiment"]
            configuration = {}
            configuration["M"] = data["params"]["hnsw_config"]["M"] 
            configuration["ef_construction"] = data["params"]["hnsw_config"]["EF_CONSTRUCTION"] 
            configuration["uploading_time"] = data["results"]["total_time"]
            configuration["used_memory"] = data["results"]["memory_usage"]["used_memory"][0]
            settings[experiment_name] = configuration
for file in files:
    if not "upload" in file:
        with open(os.path.join(path, file)) as f:
            data = json.load(f)
            result = {}
            experiment_name = data["params"]["experiment"]
            result["parallel"] = data["params"]["parallel"]
            result["ef_search"] = data["params"]["search_params"]["ef"]
            result["top"] = data["params"].get("top", 100)
            result["total_time"] = data["results"]["total_time"]
            result["mean_time"] = data["results"]["mean_time"]
            result["precision"] = data["results"]["mean_precisions"]
            result["rps"] = data["results"]["rps"]
            result["M"] = settings[experiment_name]["M"]
            result["ef_construction"] = settings[experiment_name]["ef_construction"]
            results.append(result)

result_json_path = os.path.join(path, "summary/results.json")
with open(result_json_path, "w") as f:
    json.dump(results, f)


df = pd.DataFrame(results)

# Take the highest rps from each table
idx = df.groupby(['parallel', 'ef_search', 'top', 'M', 'ef_construction'])['rps'].idxmax()
df_max_rps = df.loc[idx]

create_excel_for_top(df_max_rps, 10)
create_excel_for_top(df_max_rps, 100)