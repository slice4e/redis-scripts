import click
import json
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

def process_hnsw(results_path, summary_path):
    files = [f for f in os.listdir(results_path) if os.path.isfile(os.path.join(results_path, f))]
    settings = {}
    results = []
    for file in files:
        if "upload" in file and ".json" in file:
            with open(os.path.join(results_path, file)) as f:
                data = json.load(f)
                experiment_name = data["params"]["experiment"]
                configuration = {}
                configuration["M"] = data["params"]["hnsw_config"]["M"] 
                configuration["ef_construction"] = data["params"]["hnsw_config"]["EF_CONSTRUCTION"] 
                configuration["uploading_time"] = data["results"]["total_time"]
                configuration["used_memory"] = sum(data["results"]["memory_usage"]["used_memory"])
                settings[experiment_name] = configuration
    for file in files:
        if "search" in file and ".json" in file:
            with open(os.path.join(results_path, file)) as f:
                data = json.load(f)
                result = {}
                experiment_name = data["params"]["experiment"]
                result["data_type"] = data["params"]["search_params"].get("data_type", "FLOAT32")
                result["parallel"] = data["params"]["parallel"]
                result["ef_search"] = data["params"]["search_params"]["ef"]
                result["top"] = data["params"].get("top", 100)
                result["total_time"] = data["results"]["total_time"]
                result["mean_time"] = data["results"]["mean_time"]
                result["precision"] = data["results"]["mean_precisions"]
                result["rps"] = data["results"]["rps"]
                result["M"] = settings[experiment_name]["M"]
                result["ef_construction"] = settings[experiment_name]["ef_construction"]
                results.append(result)

    result_json_path = os.path.join(summary_path, "results.json")
    with open(result_json_path, "w") as f:
        json.dump(results, f)

    df = pd.DataFrame(results)

    # Take the highest rps from each table
    idx = df.groupby(['parallel', 'ef_search', 'data_type', 'top', 'M', 'ef_construction'])['rps'].idxmax()
    return df.loc[idx], settings

def excel_hnsw(top, df_max_rps, settings, summary_path, plot_precision):
    df_max_rps = df_max_rps[df_max_rps["top"] == top]
    df_total_time = df_max_rps.pivot(index=['parallel', 'ef_search'], columns=['data_type', 'top', 'M',  'ef_construction'], values='total_time')
    df_mean_time = df_max_rps.pivot(index=['parallel', 'ef_search'], columns=['data_type', 'top', 'M', 'ef_construction'], values='mean_time')
    df_precision = df_max_rps.pivot(index=['parallel', 'ef_search'], columns=['data_type', 'top', 'M', 'ef_construction'], values='precision')
    df_rps = df_max_rps.pivot(index=['parallel', 'ef_search'], columns=['data_type', 'top', 'M', 'ef_construction', ], values='rps')

    df_uploading_time = pd.DataFrame.from_dict(settings, orient='index')[['uploading_time']]
    df_used_memory = pd.DataFrame.from_dict(settings, orient='index')[['used_memory']]

    # Save to excel file
    excel_path = f"summary/results-hnsw-top{top}.xlsx"
    with pd.ExcelWriter(excel_path) as writer:
        df_total_time.to_excel(writer, sheet_name="Total Time")
        df_mean_time.to_excel(writer, sheet_name="Mean Time")
        df_precision.to_excel(writer, sheet_name="Precision")
        df_rps.to_excel(writer, sheet_name="RPS")
        df_uploading_time.to_excel(writer, sheet_name="Uploading Time")
        df_used_memory.to_excel(writer, sheet_name="Used Memory")

    wb = load_workbook(excel_path)
    wb.save(excel_path)

def process_svs(results_path, summary_path):
    files = [f for f in os.listdir(results_path) if (os.path.isfile(os.path.join(results_path, f))) and (f.endswith('.json'))]
    if not files:
        for root, _, filenames in os.walk(results_path):
            for filename in filenames:
                if filename.endswith('.json'):
                    files.append(os.path.join(root, filename))
    settings = {}
    results = []
    for file in files:
        if "upload" in file and ".json" in file:
            with open(os.path.join(results_path, file)) as f:
                data = json.load(f)
                if "svs" not in data["params"]["algorithm"].lower():
                    continue
                experiment_name = data["params"]["experiment"]
                configuration = {}
                configuration["num_threads"] = data["params"]["svs_tiered_config"]["NUM_THREADS"] 
                configuration["graph_degree"] = data["params"]["svs_tiered_config"]["GRAPH_DEGREE"] 
                configuration["ws_construction"] = data["params"]["svs_tiered_config"]["WS_CONSTRUCTION"] 
                configuration["quantization"] = data["params"]["svs_tiered_config"].get("QUANTIZATION", 0)
                configuration["uploading_time"] = data["results"]["total_time"]
                configuration["used_memory"] = sum(data["results"]["memory_usage"]["used_memory"])
                settings[experiment_name] = configuration

    for file in files:
        if "search" in file and ".json" in file:
            with open(os.path.join(results_path, file)) as f:
                data = json.load(f)
                if "svs" not in data["params"]["algorithm"].lower():
                    continue
                result = {}
                experiment_name = data["params"]["experiment"]
                result["parallel"] = data["params"]["parallel"]
                result["top"] = data["params"].get("top", 100)
                result["data_type"] = data["params"]["search_params"].get("data_type", "FLOAT32")
                result["total_time"] = data["results"]["total_time"]
                result["mean_time"] = data["results"]["mean_time"]
                result["precision"] = data["results"]["mean_precisions"]
                result["rps"] = data["results"]["rps"]
                result["ws_search"] = data["params"]["search_params"]["WS_SEARCH"]
                result["num_threads"] = settings[experiment_name]["num_threads"]
                result["graph_degree"] = settings[experiment_name]["graph_degree"]
                result["ws_construction"] = settings[experiment_name]["ws_construction"]
                result["quantization"] = settings[experiment_name].get("quantization", 0)
                results.append(result)

    result_json_path = os.path.join(summary_path, "results.json")
    os.makedirs(os.path.dirname(result_json_path), exist_ok=True)
    with open(result_json_path, "w") as f:
        json.dump(results, f)

    df = pd.DataFrame(results)

    idx = df.groupby(['parallel', 'ws_search', 'data_type', 'graph_degree', 'num_threads', 'top', 'quantization', 'ws_construction'])['rps'].idxmax()
    return df.loc[idx], settings

def excel_svs(top, df_max_rps, settings, summary_path, plot_precision):
    df_max_rps = df_max_rps[df_max_rps["top"] == top]
    df_total_time = df_max_rps.pivot(index=['parallel', 'ws_search'], columns=['data_type', 'graph_degree', 'num_threads', 'top', 'quantization', 'ws_construction'], values='total_time')
    df_mean_time = df_max_rps.pivot(index=['parallel', 'ws_search'], columns=['data_type', 'graph_degree', 'num_threads', 'top', 'quantization', 'ws_construction'], values='mean_time')
    df_precision = df_max_rps.pivot(index=['parallel', 'ws_search'], columns=['data_type', 'graph_degree', 'num_threads', 'top', 'quantization', 'ws_construction'], values='precision')
    df_rps = df_max_rps.pivot(index=['parallel', 'ws_search'], columns=['data_type', 'graph_degree', 'num_threads', 'top', 'quantization', 'ws_construction', ], values='rps')

    df_uploading_time = pd.DataFrame.from_dict(settings, orient='index')[['uploading_time']]
    df_used_memory = pd.DataFrame.from_dict(settings, orient='index')[['used_memory']]

    excel_path = os.path.join(summary_path, f"results-top{top}.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_total_time.to_excel(writer, sheet_name="Total Time")
        df_mean_time.to_excel(writer, sheet_name="Mean Time")
        df_precision.to_excel(writer, sheet_name="Precision")
        df_rps.to_excel(writer, sheet_name="RPS")
        df_uploading_time.to_excel(writer, sheet_name="Uploading Time")
        df_used_memory.to_excel(writer, sheet_name="Used Memory")

        if plot_precision:
            for prec in plot_precision:
                quantization_values = df_max_rps['quantization'].unique()
                parallel_values = df_max_rps['parallel'].unique()
                closest_precision_rps = []

                for quant in quantization_values:
                    for parallel in parallel_values:
                        df_quant_parallel = df_max_rps[(df_max_rps['quantization'] == quant) & (df_max_rps['parallel'] == parallel)]
                        if not df_quant_parallel.empty:
                            df_quant_parallel['precision_diff'] = abs(df_quant_parallel['precision'] - prec)
                            closest_row = df_quant_parallel.loc[df_quant_parallel['precision_diff'].idxmin()]
                            closest_precision_rps.append({
                                'quantization': quant,
                                'parallel': parallel,
                                'precision': closest_row['precision'],
                                'rps': closest_row['rps']
                            })

                df_closest_precision_rps = pd.DataFrame(closest_precision_rps)
                df_rps_pivot = df_closest_precision_rps.pivot(index='parallel', columns='quantization', values='rps')
                df_precision_pivot = df_closest_precision_rps.pivot(index='parallel', columns='quantization', values='precision')
                worksheet = writer.book.create_sheet(title=f"Precision {prec}")
                worksheet.title = f"Precision {prec}"
                
                worksheet.cell(row=1, column=1, value="RPS")
                for r in dataframe_to_rows(df_rps_pivot, index=True, header=True):
                    worksheet.append(r)
                
                start_row = df_rps_pivot.shape[0] + 5
                worksheet.cell(row=start_row, column=1, value="Precision")
                for r in dataframe_to_rows(df_precision_pivot, index=True, header=True):
                    worksheet.append(r)

                # Add Bar chart to the sheet
                chart = BarChart()
                chart.title = f"RPS @ ~{prec*100}%"
                chart.x_axis.title = "Parallel"
                chart.y_axis.title = "RPS"

                data = Reference(worksheet, min_col=2, min_row=2, max_row=df_rps_pivot.shape[0] + 3, max_col=df_rps_pivot.shape[1] + 1)
                categories = Reference(worksheet, min_col=1, min_row=2, max_row=df_rps_pivot.shape[0] + 3)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                chart.shape = 4
                worksheet.add_chart(chart, "E2")

    wb = load_workbook(excel_path)

    wb.save(excel_path)

def validate_plot_precision(ctx, param, value):
    if value:
        for v in value:
            if not (0.5 <= v <= 0.99):
                raise click.BadParameter(f"Each value in 'plot-precision' must be between 0.5 and 0.99. Invalid value: {v}")
    return value

@click.command()
@click.option('--results_path', 'results_path', type=click.Path(exists=True, file_okay=False, dir_okay=True), default='.')
@click.option('--summary_path', 'summary_path', type=click.Path(file_okay=False, dir_okay=True), default='.')
@click.option('--algorithm', 'algorithm', type=click.Choice(['SVS', 'HNSW'], case_sensitive=False), default="HNSW")
@click.option('--top', 'top', type=int, default=10)
@click.option('--plot-precision', 'plot_precision', type=float, multiple=True, callback=validate_plot_precision, help="List of precision values (0.5 to 0.99).")
def post_process(results_path, summary_path, algorithm, top, plot_precision):
    if not os.path.exists(summary_path):
        os.makedirs(summary_path)
        click.echo(f"Created summary folder: {summary_path}")

    if algorithm == "HNSW":
        df_max_rps, settings = process_hnsw(results_path, summary_path)
        excel_hnsw(top, df_max_rps, settings, summary_path, plot_precision)
    elif algorithm =="SVS":
        df_max_rps, settings = process_svs(results_path, summary_path)
        excel_svs(top, df_max_rps, settings, summary_path, plot_precision)


if __name__ == "__main__":
    post_process()